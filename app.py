from openpyxl import Workbook, load_workbook

wb = load_workbook(r'/Users/carlosinfante/Desktop/subjetcs-data.xlsx')
ws = wb.active  # give the actual worksheet
wb.create_sheet('test')

print(ws)
print(ws['A2'].value)
print(wb.sheetnames)
print(wb['test'])

# ws['A2'].value = 'Carlos'
ws['A3'].value = 'Jose'

wb.save(r'/Users/carlosinfante/Desktop/subjetcs-data.xlsx')

print(ws['A3'].value)
