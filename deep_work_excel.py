from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('tim.xlsx')
ws = wb.active
# ws.merge_cells("A1:D1")
# ws.unmerge_cells("A1:D1")
# ws.insert_rows(7)  # insert one empty row on the line 7
# ws.delete_rows(7)
# ws.insert_cols(2)
# ws.delete_cols(2)
# This moves the selected cells 2 rows bellow and 2 rows to the right. If we put minus sign then we get the contrry effect
ws.move_range("A1:D10", rows=2, cols=2)


wb.save('tim.xlsx')
